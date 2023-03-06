#!/usr/bin/python3
"""
# androguard开源库：https://github.com/androguard/androguard
# aapt工具使用介绍：https://juejin.cn/post/7075594597505695758
# aapt2工具文档：https://developer.android.google.cn/studio/command-line/aapt2?hl=zh-cn#dump_commands
# 代码参考：
https://blog.csdn.net/h1986y/article/details/123737172
https://github.com/omieo2/apk_toolbox
# 遇到编码问题，改用subprocess，参考：https://blog.csdn.net/u012871930/article/details/128022910
"""
import datetime
import getopt
import hashlib
import imghdr
import os
import re
import subprocess
import sys
import time
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from androguard.core.apk import APK

now = str(datetime.date.today())
download_dir = os.path.join(os.getcwd(), now)
imgType_list = {'jpg', 'bmp', 'png', 'jpeg', 'rgb', 'tif', 'ico'}


class Common:
    def main(self, argv):
        """
        短操作数单字母，长操作数具有语义，便于理解。长操作数可以不必全部写全但必须开头且顺序匹配的，如--version，等价于--ver,--v,--vers。前缀需要唯一
        短操作参数后带冒号需要带参加，长操作参数后面带等号需要参数
        :param argv: 命令行传递的参数，列表形式，首元素为该文件，第二个元素之后是参数
        :return:
        """
        try:
            # getopt.getopt()用于解析命令行选项参数,(args, shortopts, longopts=[])，返回两个列表，opts组成元素为短长参数及其值组成的元组，args为额外的参数
            opts, args = getopt.getopt(argv[1:], 'hvt:d:r:', ['help', 'version', 'type=', 'dirs=', 'rename='])
            andguard = AndroGuard()
            aapt = Aapt()
        except getopt.GetoptError as err:
            print('读取参数时发生错误！', err)
            self.usage()
            sys.exit()
        opts = dict(opts)
        processtype = "A"
        if '-t' in opts or '--type' in opts:
            processtype = opts.get('-t', opts.get("--type"))
        if "-h" in opts or "--help" in opts:
            self.usage()
            return
        elif "-v" in opts or "--version" in opts:
            self.version()
            return
        elif "-d" in opts or "--dirs" in opts:
            dirs = opts.get('-d', opts.get("--dirs"))
            if os.path.isdir(dirs):
                if processtype == "A":
                    andguard.app_info(dirs, True)
                elif processtype == "B":
                    aapt.app_info(dirs, True)
                else:
                    print("-t或--type的参数错误，请输入A或B")
                    return
                return
            else:
                print("-d或--dirs的参数错误，请输入一个目录路径")
                return
        elif "-r" in opts or "--rename" in opts:
            dirs = opts.get('-r', opts.get("--rename"))
            if os.path.isdir(dirs):
                self.rename(dirs)
            else:
                print("-r或--rename的参数错误，请输入一个目录路径")
                return
            return
        apk_path = " ".join(args)
        if apk_path.endswith('.apk') and os.path.isfile(apk_path):
            if processtype == "A":
                andguard.app_info(apk_path, False)
            elif processtype == "B":
                aapt.app_info(apk_path, False)
            else:
                print("-t或--type的参数错误，请输入A或B")
                return
        else:
            print("输入参数错误，请输入单个apk文件的绝对路径")
            self.usage()

    @staticmethod
    def usage():
        print(
            '''
            ———————— 使用说明 ——————————
            xxx.apk             获取apk信息
            -t, --type:         处理引擎A or B(默认A)
            -d, --dirs:         获取目录下所有apk信息
            -r, --rename:       批量格式化重命名apk
            -v, --version:      版本号
            -h, --help:         帮助信息
            —————————————————————————————
A引擎为androguard_v3.3.6-2022年11月20日，理论最高支持API29，Android10.0[https://pypi.org/project/androguard][https://github.com/androguard/androguard]
B引擎为aapt_v0.2-4913185-2018年8月10日，理论最高支持API24，Android7.0）[https://androidaapt.com],可解析A引擎获取不到图标，需要额外调用，导致apk名称不能有空格''')

    @staticmethod
    def version():
        print('———————— 版本信息 ————————')
        print('Apk_analyse v2.1.0 build on 2023年1月10日.')

    @staticmethod
    def get_file_md5(file_path):
        with open(file_path, 'rb') as f:
            md5obj = hashlib.md5()
            md5obj.update(f.read())
            md5 = md5obj.hexdigest()
            md5 = str(md5).lower()
        return md5

    @staticmethod
    def get_cert_md5(a):
        """
        获取证书md5
        :param a:
        :return:
        """
        cert_md5 = ''
        certs = set(a.get_certificates_der_v2() + [a.get_certificate_der(x) for x in a.get_signature_names()])
        for cert in certs:
            cert_md5 = hashlib.md5(cert).hexdigest()

        return cert_md5

    def rename(self, path):
        """

        :param path:
        :return:
        """
        start = time.time()
        count=0
        for root, dirs, files in os.walk(path):
            for i in range(len(files)):
                old_name = os.path.abspath(os.path.join(root, files[i]))
                if old_name.endswith(".apk"):
                    apk = APK(old_name)
                    apk_name = apk.get_app_name()
                    apk_versionname = apk.get_androidversion_name()
                    del apk
                    new_name = os.path.join(os.path.split(old_name)[0], apk_name + "_" + apk_versionname + '.apk')
                    print('%s ->> %s' % (files[i], apk_name + "_" + apk_versionname + '.apk'))
                    try:
                        os.rename(old_name, new_name)
                        count+=1
                    except FileExistsError as err:
                        print(err)
                        # 当有相同名称文件时执行去重操作
                        os.remove(old_name)
            print('\nTotal info:\n已对%d个文件进行重命名\t耗时%.2fs' % (count, time.time() - start))
            count=0


class WriteData:
    def __init__(self, filename, navigation_bar):
        self.wb = Workbook()
        # self.ws = self.wb.active，作用都是打开第一个工作表
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.filename = filename + now
        self.font = Font(name='等线', size=11)
        self.fill = PatternFill("solid", fgColor="E0EEE0")
        self.title = navigation_bar

    def autocolwid(self):
        # 第一步：计算每列最大宽度，并存储在列表lks中。
        lks = []  # 英文变量太费劲，用汉语首字拼音代替
        for i in range(1, self.ws.max_column + 1):  # 每列循环
            lk = 1  # 定义初始列宽，并在每个行循环完成后重置
            for j in range(1, self.ws.max_row + 1):  # 每行循环
                sz = self.ws.cell(row=j, column=i).value  # 每个单元格内容
                if isinstance(sz, str):  # 中文占用多个字节，需要分开处理
                    lk1 = len(sz.encode('gbk'))  # gbk解码一个中文两字节，utf-8一个中文三字节，gbk合适
                else:
                    lk1 = len(str(sz))
                if lk < lk1:
                    lk = lk1  # 借助每行循环将最大值存入lk中
                # print(lk)
            lks.append(lk)  # 将每列最大宽度加入列表。（犯了一个错，用lks = lks.append(lk)报错，append会修改列表变量，返回值none，而none不能继续用append方法）

        # 第二步：设置列宽
        for i in range(1, self.ws.max_column + 1):
            k = get_column_letter(i)  # 将数字转化为列名,26个字母以内也可以用[chr(i).upper() for i in range(97, 123)]，不用导入模块
            self.ws.column_dimensions[k].width = lks[i - 1] + 3  # 设置列宽，一般加两个字节宽度，可以根据实际情况灵活调整

    def write_data(self, row=1, data=None, is_init=False):
        """
        写入数据
        :param row:
        :param data:
        :param is_init:
        :return:
        """
        if not is_init:
            rows = self.ws.max_row
            for col in range(1, len(data) + 1):
                operate = self.ws.cell(row=rows + 1, column=col, value=data[col - 1])
                operate.font = self.font
        else:
            for i in range(1, len(self.title) + 1):
                operate = self.ws.cell(row=row, column=i, value=self.title[i - 1])
                operate.font = self.font
                operate.fill = self.fill
        self.autocolwid()
        self.wb.save(os.path.join(download_dir, self.filename + '.xlsx'))


class AndroGuard:
    def app_info(self, path, Flag):
        """

        :param path:
        :return:
        """
        start = time.time()
        if not os.path.exists(download_dir):
            os.mkdir(download_dir)

        wd = WriteData(filename='apk_info_',
                       navigation_bar=['filename', 'app_name', 'pkg_name', 'app_version', 'version_code', 'support_sdk', 'file_md5',
                                       'cert_md5'])
        wd.write_data(is_init=True)

        if Flag:
            for root, dirs, files in os.walk(path):
                for f in range(len(files)):
                    info = self.get_apk_info(files[f], root)
                    if info:
                        wd.write_data(data=info, row=f)
                    print(str(info))
                print('\nTotal info:\n已获取%d个文件信息\t耗时%.2fs' % (len(files), time.time() - start))
        else:
            root, file = os.path.split(path)
            info = self.get_apk_info(file, root)
            if info:
                wd.write_data(data=info)
            print(str(info))
            # print('参数输入有误，不是一个目录...')

    def get_apk_info(self, apkfile, root):
        """
        获取apk信息
        :param apkfile:
        :param root:
        :return:
        """
        apk_path = os.path.join(root, apkfile)
        apk_info = []
        try:
            apk = APK(apk_path)
            if apk.is_valid_APK():
                apk_name = apk.get_app_name()
                apk_info.append(apkfile)
                apk_info.append(apk_name)
                apk_info.append(apk.get_package())
                apk_info.append(apk.get_androidversion_name())
                apk_info.append(apk.get_androidversion_code())
                apk_info.append(f'{apk.get_min_sdk_version()}-{apk.get_effective_target_sdk_version()}')
                apk_info.append(Common().get_file_md5(apk_path))
                apk_info.append(Common().get_cert_md5(apk))
                icon_path = os.path.join(download_dir, f"{apk_name}_icon.png")
                icon = apk.get_file(apk.get_app_icon(max_dpi=480))
                del apk
                with open(icon_path, "wb") as f:
                    f.write(icon)
                if imghdr.what(icon_path) not in imgType_list:
                    os.remove(icon_path)
        except Exception as e:
            print(apkfile + ' ->>', e)

        return apk_info


class Aapt:
    """
    aapt 是 Android Asset Packaging Tool 的缩写，是编译和打包资源的工具，在 SDK 的 build-tools 目录下。
    Android Gradle Plugin 3.0.0 或者更高版本默认开启 aapt2（API24，Android7.0）
    """

    def app_info(self, path, Flag):
        start = time.time()
        if not os.path.exists(download_dir):
            os.mkdir(download_dir)

        wd = WriteData(filename='apk_info_',
                       navigation_bar=['filename', 'chinese_name', 'english_name', 'pkg_name', 'app_version', 'version_code', 'support_sdk',
                                       'ABIs', 'file_md5'])
        wd.write_data(is_init=True)
        if Flag:
            for root, dirs, files in os.walk(path):
                for f in range(len(files)):
                    info = self.get_apk_info(files[f], root)
                    if info:
                        wd.write_data(data=info, row=f)
                    print(str(info))
                print('\nTotal info:\n已获取%d个文件信息\t耗时%.2fs' % (len(files), time.time() - start))
        else:
            root, file = os.path.split(path)
            info = self.get_apk_info(file, root)
            if info:
                wd.write_data(data=info)
            print(str(info))

    # 检查apk版本号等信息
    def get_apk_info(self, apkfile, root):

        apk_path = os.path.join(root, apkfile)

        # query_png_command = "aapt dump --values resources %s | grep -iC10 %s" % (apk_path, pngResId)
        query_png_command = f'aapt dump --values resources {apk_path}'
        output, err = subprocess.Popen(query_png_command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.PIPE,
                                       shell=True).communicate()
        self.resource = output.decode("utf-8", "ignore")
        # 输出从 APK 的清单中提取的信息。
        get_info_command = "aapt dump badging %s" % apk_path
        output, err = subprocess.Popen(get_info_command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.PIPE,
                                       shell=True).communicate()
        output = output.decode("utf-8")
        match = re.compile("package: name='(\S+)' versionCode='(\d+)' versionName='(\S+)'").match(output)
        if not match: raise Exception("can't get packageinfo")
        packagename, versionCode, versionName = match.group(1), match.group(2), match.group(3)
        # launchableActivity = ""
        appCnName = ""
        appEnName = ""
        apkInfo = []
        xml = {}

        lines = output.splitlines()
        for l in lines:
            l = l.strip().strip("'")
            if l.startswith("application-icon"):
                iconInfo = l.split(":'")
                # 如 application-icon-160:'res/XYV.xml'
                if iconInfo[1].endswith('.xml'):
                    if iconInfo[0] not in xml: xml[iconInfo[0]] = iconInfo[1]
                # 如 application-icon-160:'r/o/ic_launcher.png'
                elif iconInfo[1].endswith('.png'):
                    self.getIconFromPng(apk_path, iconInfo)
            # if l.startswith("launchable-activity"):
            #     launchableActivity = l.split("'")[1]
            elif l.startswith("application-label-zh-CN") or l.startswith("application-label-zh"):
                appCnName = l.split(":'")[1]
            elif l.startswith("application-label-en"):
                appEnName = l.split(":'")[1]
            elif l.startswith("application-label:"):
                lableName = l.split(":'")[1]
                if appCnName == "":
                    appCnName = lableName
                if appEnName == "":
                    appEnName = lableName
            elif l.startswith("sdkVersion"):
                minsdkversion = l.split(":'")[1]
            elif l.startswith("targetSdkVersion"):
                targetSdkVersion = l.split(":'")[1]
            elif l.startswith("native-code"):
                ABIs = l.split(":")[1].replace("'", "").strip()

        self.getIconFromXml(apk_path, xml)
        file_md5 = Common().get_file_md5(apk_path)
        support_sdk = f"{minsdkversion}-{targetSdkVersion}"
        # apkInfo = f"filename：{filename}\nchinese_name：{appCnName}\nenglish_name：{appEnName}\npkg_name：{packagename}\napp_version：{versionName}\nversion_code：{versionCode}\nsupport_sdk：{support_sdk}\nABIs：{ABIs} \nfile_md5：{file_md5}:

        apkInfo = [apkfile, appCnName, appEnName, packagename, versionName, versionCode, support_sdk, ABIs, file_md5]
        # 将apk的信息写入文本文件中
        return apkInfo

    # 从apk中提取出PNG图片
    def getIconFromPng(self, apk_path, iconInfo):
        z = zipfile.ZipFile(apk_path, 'r')
        icon = z.read(iconInfo[1])
        with open('%s/%s.png' % (download_dir, iconInfo[0]), 'wb') as f:
            f.write(icon)

    # 通过aapt查询图片
    def getIconFromXml(self, apk_path, xml):
        png = []
        v = []
        for key, value in xml.items():
            if value in v:
                continue
            else:
                iconInfo = [key, value]
                v.append(value)
                # xmltree 获取指定 APK 的指定xml文件 如：aapt dump xmltree C:\Users\Administrator\Desktop\C.apk res/XYV.xml
                query_xml_command = "aapt dump xmltree %s %s" % (apk_path, iconInfo[1])
                output = os.popen(query_xml_command)
                # 解决os.popen()读取乱码问题：https://blog.csdn.net/wenxingchen/article/details/119649289
                output = output.buffer.read().decode('utf-8')
                lines = output.splitlines()
                """N: android = http://schemas.android.com/apk/res/android
                E: adaptive - icon(line=2)
                E: background(line=3)
                A: android:drawable(0x01010199) = @0x7f08022a
                E: foreground(line=4)
                A: android:drawable(0x01010199) =@0x7f080233
                E: monochrome(line=5)
                A: android:drawable(0x01010199) = @0x7f080233"""
                # 图片资源ID根据arsc文件转具体图片路径
                pngResId = ''
                for l in lines:
                    l = l.strip()
                    if (l.startswith('A: android:drawable')):
                        pngResId = l.split("@0x")[1]
                    if not pngResId:
                        continue
                    """7f08022a
                    7f080233
                    7f080233"""
                    pattern = re.compile(pngResId + r'.{1,200}?\(string8\)(.{1,50}?png")', re.S).findall(self.resource)
                    if not pattern: continue
                    [png.append(i.strip(' ').strip('"')) for i in pattern if i not in png]
        png = list(zip(xml.keys(), png))
        for pngPath in png: self.getIconFromPng(apk_path, pngPath)


if __name__ == '__main__':
    # 获取命令行传递的参数
    Common().main(sys.argv)
